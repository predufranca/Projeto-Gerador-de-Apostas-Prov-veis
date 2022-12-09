# Projeto Peaky Blinders

from openpyxl import load_workbook
import random

from scipy import rand

list = []
caminho = 'PythonBasico\projetinfellas\mega-sena.xlsx'
arquivo_excel = load_workbook(caminho)
planilha = arquivo_excel.active

max_linha = planilha.max_row
max_col = planilha.max_column

cell = planilha.cell(row=1, column=1).value
 
if not isinstance(cell, int):

    for num in range(1, 8):
        planilha.delete_rows(num)

    planilha.delete_cols(1, 2)

    planilha.delete_rows(1, 3)

    arquivo_excel.save('PythonBasico\projetinfellas\mega-sena.xlsx')


for linha in range(1, max_linha+1):
    for coluna in range(1, max_col+1):
        list.append(planilha.cell(row=linha, column=coluna).value)
        list.sort()

num = 1
result = [
    # Dezenas
    [*range(61)],
    # Quantidade 
    [*range(61)]
]

for value in list:
    
    while num <= len(list):

        if num == value:
            result[1][num] = 0
            result[1][num] = list.count(num)
            break
        else:
            result[1][num] = 0
            result[1][num] = list.count(num)
            print(f'O número {num} se repetiu {result[1][num]} vez(es).')
            num += 1
            break

print(f'O número {num} se repetiu {result[1][num]} vez(es).')

# Fazer laço que verifica qual número se repete mais num intervalo (que inclusive você tem que pensar em qual vai serkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk)
#teviraaimalandro
#faleitoleve

probally_bets = [*range(19)]
option_index = []
index = 0

for num in probally_bets:
    
    probally_bets[num] = max(result[1])
    index = result[1].index(max(result[1]))
    option_index.append(result[0][index])
    del(result[0][index])
    del(result[1][index])

finally_bets = []

repeat = input("\nCom base nos dados acima, digite a quantidade de jogos que quer fazer: ")

if repeat.isnumeric():

    repeat = int(repeat)

    for i in range(repeat):

        for index in range(6):
            finally_bets.append(option_index[random.randrange(0, len(option_index)-1)])
            option_index.remove(finally_bets[index])
            finally_bets.sort()
        
        print(f'\n{i+1}º Jogo: {finally_bets}')

        option_index.extend(finally_bets)
        finally_bets.clear()

else:
    print(f'{repeat} não é um número, seu animal. Mais sorte na próxima.')
    exit()
